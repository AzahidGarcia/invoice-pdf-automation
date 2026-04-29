"""Tests for the Excel summary module."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from invoice_automation.excel_summary import ExcelSummaryError, generate_excel_summary
from invoice_automation.models import Invoice, InvoiceItem


class TestExcelSummary:
    """Tests for generate_excel_summary function."""

    def test_generate_excel(self, sample_invoice: Invoice, tmp_path: Path) -> None:
        """Test that Excel is generated with correct sheets."""
        output_path = tmp_path / "summary.xlsx"
        result = generate_excel_summary([sample_invoice], output_path)

        assert result == output_path
        assert output_path.exists()
        assert output_path.stat().st_size > 100

    def test_excel_has_four_sheets(self, sample_invoice: Invoice, tmp_path: Path) -> None:
        """Test that Excel has all 4 required sheets."""
        output_path = tmp_path / "summary.xlsx"
        generate_excel_summary([sample_invoice], output_path)

        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames
        assert "Resumen general" in sheet_names
        assert "Por cliente" in sheet_names
        assert "Por mes" in sheet_names
        assert "Por categoría" in sheet_names

    def test_empty_list_raises(self, tmp_path: Path) -> None:
        """Test that empty invoice list raises ExcelSummaryError."""
        output_path = tmp_path / "summary.xlsx"
        with pytest.raises(ExcelSummaryError, match="No invoices"):
            generate_excel_summary([], output_path)

    def test_multiple_invoices_aggregated(
        self, sample_invoice: Invoice, tmp_path: Path
    ) -> None:
        """Test that multiple invoices are properly aggregated."""
        # Create second invoice
        invoice2 = Invoice(
            invoice_id="TEST-002",
            client_name="Cliente dos",
            client_rfc="XAXX010101002",
            issue_date=__import__("datetime").date(2026, 4, 2),
            items=[InvoiceItem(description="Item", quantity=1, unit_price=200.00)],
            category="Electronica",
        )

        output_path = tmp_path / "summary.xlsx"
        generate_excel_summary([sample_invoice, invoice2], output_path)

        wb = load_workbook(output_path)
        # Check "Por cliente" sheet has both clients
        client_sheet = wb["Por cliente"]
        client_names = [cell.value for cell in list(client_sheet.columns)[0]]
        assert "Cliente de prueba" in client_names
        assert "Cliente dos" in client_names
