"""Excel summary generation with OpenPyXL."""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from invoice_automation.models import Invoice


class ExcelSummaryError(Exception):
    """Raised when Excel generation fails."""


HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ALT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_excel_summary(invoices: list[Invoice], output_path: str | Path) -> Path:
    """
    Generate an Excel workbook with multiple summary sheets.

    Args:
        invoices: List of validated invoices
        output_path: Where to save the Excel file

    Returns:
        Path to the generated Excel file
    """
    path = Path(output_path)
    if not invoices:
        raise ExcelSummaryError("No invoices to summarize")

    try:
        wb = Workbook()
    except Exception as e:
        raise ExcelSummaryError(f"Failed to create workbook: {e}") from e

    # Remove default sheet
    if wb.active is not None:
        wb.remove(wb.active)

    # Sheet 1: Resumen general
    _add_resumen_general(wb, invoices)

    # Sheet 2: Por cliente
    _add_por_cliente(wb, invoices)

    # Sheet 3: Por mes
    _add_por_mes(wb, invoices)

    # Sheet 4: Por categoría
    _add_por_categoria(wb, invoices)

    try:
        wb.save(str(path))
    except Exception as e:
        raise ExcelSummaryError(f"Failed to save workbook: {e}") from e

    return path


def _add_resumen_general(wb: Workbook, invoices: list[Invoice]) -> None:
    """Add the general summary sheet."""
    ws = wb.create_sheet("Resumen general")

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Resumen General de Ventas"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center")

    total_subtotal = sum(float(i.subtotal) for i in invoices)
    total_iva = sum(float(i.iva_amount) for i in invoices)
    total_total = sum(float(i.total) for i in invoices)

    headers: list[str] = ["Métrica", "Valor"]
    row1: list[str] = ["Total de facturas", str(len(invoices))]
    row2: list[str] = ["Subtotal", f"${total_subtotal:,.2f}"]
    row3: list[str] = ["IVA 16%", f"${total_iva:,.2f}"]
    row4: list[str] = ["Total", f"${total_total:,.2f}"]
    data: list[list[str]] = [headers, row1, row2, row3, row4]

    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            if row_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            elif row_idx == len(data) + 2:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def _add_por_cliente(wb: Workbook, invoices: list[Invoice]) -> None:
    """Add sales by client sheet."""
    ws = wb.create_sheet("Por cliente")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Ventas por Cliente"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Aggregate by client
    client_data: dict[str, dict[str, Any]] = {}
    for inv in invoices:
        if inv.client_name not in client_data:
            client_data[inv.client_name] = {
                "facturas": 0,
                "subtotal": 0.0,
                "iva": 0.0,
                "total": 0.0,
            }
        client_data[inv.client_name]["facturas"] += 1
        client_data[inv.client_name]["subtotal"] += float(inv.subtotal)
        client_data[inv.client_name]["iva"] += float(inv.iva_amount)
        client_data[inv.client_name]["total"] += float(inv.total)

    headers = ["Cliente", "Facturas", "Subtotal", "IVA", "Total"]
    data = [headers]
    for client, vals in sorted(client_data.items()):
        data.append(
            [
                client,
                vals["facturas"],
                f"${vals['subtotal']:,.2f}",
                f"${vals['iva']:,.2f}",
                f"${vals['total']:,.2f}",
            ]
        )

    # Totals row
    total_facturas = sum(v["facturas"] for v in client_data.values())
    total_subtotal = sum(v["subtotal"] for v in client_data.values())
    total_iva = sum(v["iva"] for v in client_data.values())
    total_total = sum(v["total"] for v in client_data.values())
    data.append(["TOTAL", total_facturas, f"${total_subtotal:,.2f}", f"${total_iva:,.2f}", f"${total_total:,.2f}"])

    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            if row_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            elif row_idx == len(data) + 2:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18


def _add_por_mes(wb: Workbook, invoices: list[Invoice]) -> None:
    """Add monthly sales sheet."""
    ws = wb.create_sheet("Por mes")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Ventas por Mes"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Aggregate by month
    month_data: dict[str, dict[str, Any]] = {}
    for inv in invoices:
        month_key = inv.issue_date.strftime("%Y-%m")
        if month_key not in month_data:
            month_data[month_key] = {
                "facturas": 0,
                "subtotal": 0.0,
                "iva": 0.0,
                "total": 0.0,
            }
        month_data[month_key]["facturas"] += 1
        month_data[month_key]["subtotal"] += float(inv.subtotal)
        month_data[month_key]["iva"] += float(inv.iva_amount)
        month_data[month_key]["total"] += float(inv.total)

    headers = ["Mes", "Facturas", "Subtotal", "IVA", "Total"]
    data = [headers]
    for month, vals in sorted(month_data.items()):
        month_name = pd.to_datetime(month + "-01").strftime("%B %Y")
        data.append(
            [
                month_name,
                vals["facturas"],
                f"${vals['subtotal']:,.2f}",
                f"${vals['iva']:,.2f}",
                f"${vals['total']:,.2f}",
            ]
        )

    # Totals row
    total_facturas = sum(v["facturas"] for v in month_data.values())
    total_subtotal = sum(v["subtotal"] for v in month_data.values())
    total_iva = sum(v["iva"] for v in month_data.values())
    total_total = sum(v["total"] for v in month_data.values())
    data.append(["TOTAL", total_facturas, f"${total_subtotal:,.2f}", f"${total_iva:,.2f}", f"${total_total:,.2f}"])

    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            if row_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            elif row_idx == len(data) + 2:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18


def _add_por_categoria(wb: Workbook, invoices: list[Invoice]) -> None:
    """Add sales by category sheet."""
    ws = wb.create_sheet("Por categoría")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Ventas por Categoría"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Aggregate by category
    cat_data: dict[str, dict[str, Any]] = {}
    for inv in invoices:
        if inv.category not in cat_data:
            cat_data[inv.category] = {
                "facturas": 0,
                "subtotal": 0.0,
                "iva": 0.0,
                "total": 0.0,
            }
        cat_data[inv.category]["facturas"] += 1
        cat_data[inv.category]["subtotal"] += float(inv.subtotal)
        cat_data[inv.category]["iva"] += float(inv.iva_amount)
        cat_data[inv.category]["total"] += float(inv.total)

    headers = ["Categoría", "Facturas", "Subtotal", "IVA", "Total"]
    data = [headers]
    for cat, vals in sorted(cat_data.items()):
        data.append(
            [
                cat,
                vals["facturas"],
                f"${vals['subtotal']:,.2f}",
                f"${vals['iva']:,.2f}",
                f"${vals['total']:,.2f}",
            ]
        )

    # Totals row
    total_facturas = sum(v["facturas"] for v in cat_data.values())
    total_subtotal = sum(v["subtotal"] for v in cat_data.values())
    total_iva = sum(v["iva"] for v in cat_data.values())
    total_total = sum(v["total"] for v in cat_data.values())
    data.append(["TOTAL", total_facturas, f"${total_subtotal:,.2f}", f"${total_iva:,.2f}", f"${total_total:,.2f}"])

    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            if row_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            elif row_idx == len(data) + 2:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
